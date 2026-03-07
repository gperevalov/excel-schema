from openpyxl.worksheet.worksheet import Worksheet

from excel_schema_engine import SheetSchema
from excel_schema_engine.global_vars import Language, ValidatorErrComment
from excel_schema_engine.utils import flatten_columns


class ExcelValidator:
    """Validate worksheet headers against a SheetSchema definition."""

    def __init__(self, schema: SheetSchema, language: Language = Language.EN):
        """Initialize validator with a sheet schema and target language."""
        self.schema = schema
        self.i18n = ValidatorErrComment(language)

    def validate_headers(self, ws: Worksheet) -> list[str]:
        """Return a list of human-readable header validation errors."""
        expected = [
            column.header
            for column in flatten_columns(self.schema.columns)
        ]

        found = []

        for col in range(1, ws.max_column + 1):
            sub = ws.cell(row=2, column=col).value
            head = ws.cell(row=1, column=col).value

            value = sub if sub else head

            found.append(str(value).strip() if value else None)

        errors: list[str] = []

        # проверка количества колонок
        if len(found) < len(expected):

            errors.append(
                self.i18n.get(
                    "columns_count",
                    expected=len(expected),
                    found=len(found)
                )
            )

        for idx, header in enumerate(expected):

            if idx >= len(found):

                errors.append(
                    self.i18n.get(
                        "missing_column",
                        column=header
                    )
                )
                continue

            if found[idx] != header:

                errors.append(
                    self.i18n.get(
                        "wrong_header",
                        index=idx + 1,
                        expected=header,
                        found=found[idx],
                    )
                )

        return errors