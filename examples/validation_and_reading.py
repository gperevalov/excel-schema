from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from excel_schema_engine import (
    ExcelSchema,
    SheetSchema,
    Column,
    CellStyle,
    ExcelReader,
    ExcelValidator,
)
from excel_schema_engine.global_vars import Language


def build_schema() -> ExcelSchema:
    """Return a schema that matches the example 'Products' worksheet."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    styles = {
        "header": CellStyle(
            font=Font(bold=True, color="FFFFFF", size=11),
            fill=PatternFill("solid", fgColor="000066CC"),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=thin_border,
        ),
    }

    return ExcelSchema(
        author="Your Name",
        styles=styles,
        sheets=[
            SheetSchema(
                name="Products",
                columns=[
                    Column(key="sku", header="SKU", style="header"),
                    Column(key="offer", header="Offer", style="header"),
                    Column(
                        key="strategy",
                        header="Strategy",
                        style="header",
                        subcolumns=[
                            Column(key="delta_type", header="Type", style="header"),
                            Column(key="delta", header="Delta", style="header"),
                        ],
                    ),
                ],
            ),
        ],
    )


def validate_and_read(path: str = "./examples/products.xlsx") -> None:
    """Validate headers and print all data rows from the given workbook."""
    schema = build_schema()

    wb = load_workbook(path)
    ws = wb["Products"]

    # 1) Validate headers
    validator = ExcelValidator(schema.sheets[0], language=Language.EN)
    errors = validator.validate_headers(ws)

    if errors:
        print("Header errors:")
        for err in errors:
            print("-", err)
        return

    # 2) Read all rows
    reader = ExcelReader(schema.sheets[0])

    for row in reader.iter_rows(ws, start_row=3):
        print(row)


if __name__ == "__main__":
    validate_and_read()
