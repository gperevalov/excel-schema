from dataclasses import dataclass, field
from typing import List, Dict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


"""
Object examples:

thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

styles = {
    "header_green": CellStyle(
        font=Font(bold=True, color='000000', size=11),
        fill=PatternFill("solid", fgColor="08ff31"),
        alignment=Alignment(horizontal='center', vertical='center'),
        border=thin_border
    ),
    "header_blue": CellStyle(
        font=Font(bold=True, color='FFFFFF', size=11),
        fill=PatternFill("solid", fgColor="000066CC"),
        alignment=Alignment(horizontal='center', vertical='center'),
        border=thin_border
    )
}

schema = ExcelSchema(
        author="pytest",
        styles=styles,
        sheets=[

            SheetSchema(
                name="Products",
                columns=[

                    Column(
                        key="sku",
                        header="SKU",
                        style="header_green"
                    ),

                    Column(
                        key="offer",
                        header="Offer",
                        style="header_blue",
                        comment=Comment(
                            comment = "Comment 1"
                            width = 100
                            height = 100
                        )
                    ),

                    Column(
                        key="strategy",
                        header="Strategy",
                        style="header_green",
                        subcolumns=[

                            Column(
                                key="delta_type",
                                header="Type",
                                style="header_blue",
                                comment=Comment(
                                    comment = "Comment 2"
                                    width = 100
                                    height = 100
                                )
                            ),

                            Column(
                                key="delta",
                                header="Delta",
                                style="header_green",
                                comment=Comment(
                                    comment = "Comment 3"
                                    width = 100
                                    height = 100
                                )
                            ),

                        ]
                    ),

                ]
            )

        ]
    )
"""


@dataclass
class Comment:
    """Configuration for a cell comment attached to a header cell."""
    comment: str
    height: int = 144
    width: int = 79


@dataclass
class Column:
    """Column definition used in a SheetSchema (optionally with nested subcolumns)."""
    key: str # variable name
    header: str # header text
    comment: Comment | None = None
    style: str | None = None # cell style
    subcolumns: list["Column"] = field(default_factory=list) # list cell under header cell


@dataclass
class SheetSchema:
    """Schema for a single worksheet including its name and top-level columns."""
    name: str # sheet name
    columns: List[Column] # list columns


@dataclass
class CellStyle:
    """Wrapper around openpyxl style objects used in schema definitions."""
    font: Font | None = None # for openpyxl.styles.Font
    fill: PatternFill | None = None # for openpyxl.styles.PatternFill
    alignment: Alignment | None = None # for openpyxl.styles.Alignment
    border: Border | None = None # for openpyxl.styles.Border


@dataclass
class ExcelSchema:
    """Top-level Excel schema combining sheets, shared styles and author metadata."""
    sheets: List[SheetSchema] # list sheets with columns
    styles: Dict[str, CellStyle] = field(default_factory=dict)
    author: str = '' # author for comments


@dataclass
class ExcelErrorsSchema:
    """Styles used by ExcelErrors to mark error, fixed and highlighted cells."""
    error_fill: CellStyle = field(default_factory=lambda: CellStyle(
        font=Font(bold=True, color='FFFFFF', size=11),
        fill=PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
        alignment=Alignment(horizontal='center', vertical='center'),
        border=Border(
            left=Side(border_style="thin", color="808080"),
            right=Side(border_style="thin", color="808080"),
            top=Side(border_style="thin", color="808080"),
            bottom=Side(border_style="thin", color="808080")
        )
    ))
    fixed_fill: CellStyle = field(default_factory=lambda:  CellStyle(
        font=Font(bold=True, color='000000', size=11),
        fill=PatternFill(start_color="78FF66", end_color="78FF66", fill_type="solid"),
        alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
        border=Border(
            left=Side(border_style="thin", color="808080"),
            right=Side(border_style="thin", color="808080"),
            top=Side(border_style="thin", color="808080"),
            bottom=Side(border_style="thin", color="808080")
        )
    ))
    highlight_fill: CellStyle = field(default_factory=lambda:  CellStyle(
        fill=PatternFill(start_color="FFB979", end_color="FFB979", fill_type="solid"),
        border=Border(
            left=Side(border_style="thin", color="808080"),
            right=Side(border_style="thin", color="808080"),
            top=Side(border_style="thin", color="808080"),
            bottom=Side(border_style="thin", color="808080")
        )
    ))
    author: str = ''
