from typing import Iterator

from openpyxl.worksheet.worksheet import Worksheet

from excel_schema_engine import SheetSchema
from excel_schema_engine.utils import flatten_columns


class ExcelReader:
    """Read worksheet rows into dicts using a SheetSchema column definition."""

    def __init__(self, sheet_schema: SheetSchema):
        """Initialize reader with the sheet schema that describes the columns."""
        self.sheet_schema = sheet_schema
        self.column_map: dict[str, int] = self._build_column_map()

    def _build_column_map(self) -> dict[str, int]:
        """Build a mapping from column keys to 1-based column indexes."""
        return {
            column.key: idx + 1
            for idx, column in enumerate(
                flatten_columns(self.sheet_schema.columns)
            )
        }

    def parse_row(self, row: tuple) -> dict:
        """Convert a raw tuple row into a dict keyed by column keys."""
        return {
            key: row[col - 1]
            for key, col in self.column_map.items()
        }

    def iter_rows(
        self,
        ws: Worksheet,
        start_row: int = 3
    ) -> Iterator[dict]:
        """Yield parsed dict rows starting from the given row index."""
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            yield self.parse_row(row)

    def read_all(self, ws: Worksheet, start_row: int = 3) -> list[dict]:
        """Read all rows from the worksheet into a list of dicts."""
        return list(self.iter_rows(ws, start_row))