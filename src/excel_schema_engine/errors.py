from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from excel_schema_engine import CellStyle
from excel_schema_engine.global_vars import ValidatorErrComment, Language


class ExcelErrors:
    """Helpers for marking, highlighting and updating error cells in a worksheet."""

    def __init__(self, error_schema, language: Language = Language.EN):
        """Initialize with an ExcelErrorsSchema and target language for messages."""
        self.error_schema = error_schema
        self.error_prefix = ValidatorErrComment(language).get("error_prefix")
        self.warning_prefix = ValidatorErrComment(language).get("warning_prefix")

    def mark_error(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        msg: str | None = None,
        height: int = 79,
        width: int = 144,
        custom_fill: CellStyle = None,
    ):
        """Mark cell as error"""
        cell = ws.cell(row=row, column=col)

        if custom_fill is not None:
            fill = custom_fill
        else:
            fill = self.error_schema.error_fill

        self._apply_style(cell, fill)

        if msg is not None:
            cell.comment = Comment(
                self.error_prefix + msg,
                self.error_schema.author,
                height,
                width
            )

    def mark_warning(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        msg: str | None = None,
        height: int = 79,
        width: int = 144,
        custom_fill: CellStyle = None,
    ):
        """
        Mark cell as warning
        use me like 'def mark_error()'
        """
        cell = ws.cell(row=row, column=col)

        if custom_fill is not None:
            fill = custom_fill
        else:
            fill = self.error_schema.warning_fill

        self._apply_style(cell, fill)

        if msg is not None:
            cell.comment = Comment(
                self.warning_prefix + msg,
                self.error_schema.author,
                height,
                width
            )

    def highlight_row(self, ws: Worksheet, row: int, custom_fill: CellStyle = None):
        """Highlight all cells in a row using the highlight or a custom fill style."""
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if custom_fill is not None:
                fill = custom_fill
            else:
                fill = self.error_schema.highlight_fill
            self._apply_style(cell, fill)

    def mark_fixed(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        msg: str | None = None,
        height: int = 79,
        width: int = 144,
        custom_fill: CellStyle = None,
    ):
        """Mark cell as fixed"""
        cell = ws.cell(row=row, column=col)

        if cell.comment is None:
            return

        if self.error_prefix.lower() not in cell.comment.text.lower():
            return

        if custom_fill is not None:
            fill = custom_fill
        else:
            fill = self.error_schema.fixed_fill

        self._apply_style(cell, fill)

        if msg:
            cell.comment = Comment(
                msg,
                self.error_schema.author,
                height,
                width
            )

    @staticmethod
    def _apply_style(cell, style):
        """Apply the given CellStyle to a single cell."""
        if style.font:
            cell.font = style.font

        if style.fill:
            cell.fill = style.fill

        if style.alignment:
            cell.alignment = style.alignment

        if style.border:
            cell.border = style.border
