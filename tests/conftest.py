import pytest
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from excel_schema_engine import (
    ExcelSchema,
    SheetSchema,
    Column,
    CellStyle
)


@pytest.fixture
def schema():

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    styles = {
        "header_green": CellStyle(
            font=Font(bold=True, color='000000', size=11),
            fill=PatternFill("solid", fgColor="08ff31"),
            alignment=Alignment(horizontal='center', vertical='center'),
            border=thin_border
        ),
        "header_blue": CellStyle(
            font=Font(bold=True, color='FFFFFF', size=11),
            fill=PatternFill("solid", fgColor="000066CC"),
            alignment=Alignment(horizontal='center', vertical='center'),
            border=thin_border
        )
    }

    return ExcelSchema(
        author="pytest",
        styles=styles,
        sheets=[

            SheetSchema(
                name="Products",
                columns=[

                    Column(
                        key="sku",
                        header="SKU",
                        style="header_green"
                    ),

                    Column(
                        key="offer",
                        header="Offer",
                        style="header_blue",
                        comment="offer_id"
                    ),

                    Column(
                        key="strategy",
                        header="Strategy",
                        style="header_green",
                        subcolumns=[

                            Column(
                                key="delta_type",
                                header="Type",
                                style="header_blue",
                                comment="Very blue"
                            ),

                            Column(
                                key="delta",
                                header="Delta",
                                style="header_green",
                                comment="Very green"
                            ),

                        ]
                    ),

                ]
            )

        ]
    )