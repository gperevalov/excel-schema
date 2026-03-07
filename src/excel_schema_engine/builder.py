from openpyxl import Workbook
from openpyxl.comments import Comment

from excel_schema_engine.global_vars import Language
from excel_schema_engine.utils import flatten_columns
from excel_schema_engine.global_vars import ValidatorErrComment


class ExcelBuilder:

    def __init__(self, schema):
        self.schema = schema

    def build(self):
        """generate headers cell"""
        wb = Workbook()

        for i, sheet_schema in enumerate(self.schema.sheets):

            ws = wb.active if i == 0 else wb.create_sheet(sheet_schema.name)
            ws.title = sheet_schema.name

            self._build_headers(ws, sheet_schema)

        return wb

    def _build_headers(self, ws, sheet_schema):

        col = 1

        for column in sheet_schema.columns:

            if column.subcolumns:

                ws.merge_cells(
                    start_row=1,
                    start_column=col,
                    end_row=1,
                    end_column=col + len(column.subcolumns) - 1
                )

                cell = ws.cell(row=1, column=col, value=column.header)

                self._apply_style(cell, column.style)
                self._apply_comment(cell, column.comment)

                for i, sub in enumerate(column.subcolumns):

                    sub_cell = ws.cell(
                        row=2,
                        column=col + i,
                        value=sub.header
                    )

                    self._apply_style(sub_cell, sub.style)
                    self._apply_comment(sub_cell, sub.comment)

                col += len(column.subcolumns)

            else:

                ws.merge_cells(
                    start_row=1,
                    start_column=col,
                    end_row=2,
                    end_column=col
                )

                cell = ws.cell(row=1, column=col, value=column.header)

                self._apply_style(cell, column.style)
                self._apply_comment(cell, column.comment)

                col += 1

    def _apply_style(self, cell, style_name):

        if not style_name:
            return

        style = self.schema.styles.get(style_name)

        if not style:
            return

        if style.font:
            cell.font = style.font

        if style.fill:
            cell.fill = style.fill

        if style.alignment:
            cell.alignment = style.alignment

        if style.border:
            cell.border = style.border

    def _apply_comment(self, cell, comment):

        if comment:
            cell.comment = Comment(comment.comment, self.schema.author, comment.height, comment.width)

    def write_rows(self, ws, sheet_name, rows, start_row=3, language: Language = Language.EN):
        i18n = ValidatorErrComment(language)
        sheet_schema = None
        for sheet in self.schema.sheets:
            if sheet.name == sheet_name:
                sheet_schema = sheet

        if not sheet_schema:
            print(
                i18n.get(
                    "miss_sheet",
                    sheet_name=sheet_name,
                )
            )
            return

        flat_columns = flatten_columns(sheet_schema.columns)

        for row_idx, data in enumerate(rows, start=start_row):

            for col_idx, column in enumerate(flat_columns, start=1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=data.get(column.key)
                )
